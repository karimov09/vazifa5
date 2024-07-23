import os
from openpyxl import load_workbook
class LoadExcel:
    def __init__(self, filename):
        self.filename = filename
        if not os.path.isfile(self.filename):
            raise FileNotFoundError(f" {self.filename}")
    def reader(self):
        try:
            book = load_workbook(filename=self.filename)
            sheet = book.active
            for row in sheet.iter_rows(values_only=True):
                print(row)
        except Exception as e:
            print(f"xato:{Exception}")
n_name = 'user.xlsx'
try:
    loader = LoadExcel(n_name)
    loader.reader()
except FileNotFoundError as e:
    print(e)